"""导出工具 — 输出 CSV / Excel"""

import os
import csv
from typing import Optional

from config import REPORTS_DIR
from product import Product


def export_csv(products: list[Product], filename: Optional[str] = None) -> str:
    """导出为 CSV 文件

    Returns:
        导出文件路径
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    if not filename:
        filename = "shopee_products.csv"

    path = os.path.join(REPORTS_DIR, filename)

    fieldnames = [
        "item_id", "title", "price_min", "price_max", "currency",
        "historical_sold", "rating", "review_count",
        "shop_id", "shop_name", "shop_location",
        "liked_count", "is_preferred_plus",
    ]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for p in products:
            # 只写入 CSV 兼容的字段
            row = {k: getattr(p, k, "") for k in fieldnames}
            writer.writerow(row)

    return path


def export_excel(products: list[Product], filename: Optional[str] = None) -> Optional[str]:
    """导出为 Excel 文件（需要 openpyxl）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  ⚠️ openpyxl 未安装，回退到 CSV")
        return export_csv(products, filename)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    if not filename:
        filename = "shopee_products.xlsx"
    elif not filename.endswith(".xlsx"):
        filename += ".xlsx"

    path = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "选品数据"

    # 表头
    headers = [
        "商品ID", "标题", "最低价(TWD)", "最高价(TWD)",
        "月销", "评分", "评价数",
        "店铺ID", "店铺名", "店铺位置",
        "收藏数", "优选+",
    ]

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, p in enumerate(products, 2):
        data = [
            p.item_id,
            p.title,
            p.price_min,
            p.price_max,
            p.historical_sold,
            p.rating,
            p.review_count,
            p.shop_id,
            p.shop_name,
            p.shop_location,
            p.liked_count,
            "是" if p.is_preferred_plus else "否",
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # 列宽
    col_widths = [12, 50, 12, 12, 8, 6, 8, 12, 20, 12, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(path)
    return path
