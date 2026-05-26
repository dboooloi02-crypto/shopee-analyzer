"""
Shopee 选品数据分析工具 — 第一版 (手动输入版)

当前 Shopee 反爬严格，第一版先支持：
  1. 手动数据输入（用户粘贴 CSV）
  2. 1688 价格爬虫（辅助比价）
  3. 自动分析 + Excel 输出

使用方式：
  python3 main.py                          # 使用示例数据演示
  python3 main.py data/parsed/sample.json  # 使用已有数据文件
"""

import sys
import os
import csv
import json
import re
from datetime import datetime
from dataclasses import dataclass, asdict


def disp_width(s: str) -> int:
    """计算字符串在终端中的显示宽度（中文=2，英文=1）"""
    width = 0
    for ch in s:
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            width += 2
        else:
            width += 1
    return width


def pad_disp(s: str, width: int) -> str:
    """按显示宽度填充空格对齐"""
    return s + ' ' * max(0, width - disp_width(s))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
REPORTS_DIR = os.path.join(DATA_DIR, "reports")
SAMPLE_DIR = os.path.join(DATA_DIR, "parsed")


# ============================================================
# 数据结构
# ============================================================

@dataclass
class Product:
    title: str
    price_twd: float          # Shopee 售价（TWD）
    price_cny: float = 0      # 对应 1688 批发价（CNY）
    historical_sold: int = 0
    rating: float = 0.0
    review_count: int = 0
    shop_name: str = ""
    shop_location: str = ""
    item_id: str = ""
    # 计算字段
    platform_fee_pct: float = 0.15
    shipping_twd: float = 60  # 台湾岛内物流
    def gross_profit(self) -> float:
        """毛利润 (TWD)"""
        revenue = self.price_twd * 0.95  # 扣平台抽成5%
        cost_cny = self.price_cny
        cost_twd = cost_cny * 4.5       # 假设汇率 1 CNY = 4.5 TWD
        shipping_tw = self.shipping_twd
        shipping_cn = 15                # 大陆到台湾物流约15 RMB
        total_cost = cost_twd + shipping_tw + shipping_cn * 4.5
        return revenue - total_cost
    
    def profit_margin_pct(self) -> float:
        """毛利率"""
        if self.price_twd <= 0:
            return 0
        return round(self.gross_profit() / self.price_twd * 100, 1)


# ============================================================
# 分析引擎
# ============================================================

def analyze_products(products: list[Product]) -> list[dict]:
    """对商品列表进行分析，返回增强数据"""
    results = []
    for p in products:
        row = {
            "标题": p.title,
            "售价(TWD)": p.price_twd,
            "批发价(CNY)": p.price_cny,
            "月销": p.historical_sold,
            "评分": p.rating,
            "评价数": p.review_count,
            "毛利率(%)": p.profit_margin_pct(),
            "毛利(TWD)": round(p.gross_profit(), 1),
            "风险等级": "",
            "建议操作": "",
        }
        # 风险评分
        risks = []
        if p.profit_margin_pct() < 10:
            risks.append("利润过低")
        if 10 <= p.profit_margin_pct() < 20:
            risks.append("利润偏低")
        if p.rating < 4.0 and p.review_count > 10:
            risks.append("评分差")
        if p.historical_sold < 5 and p.review_count == 0:
            risks.append("无数据参考")
        
        row["风险等级"] = "高" if (p.profit_margin_pct() < 10 or len(risks) >= 2) else ("中" if len(risks) >= 1 else "低")
        row["风险提示"] = " | ".join(risks) if risks else "正常"
        
        # 建议
        if row["风险等级"] == "高":
            row["建议操作"] = "❌ 不推荐"
        elif p.historical_sold >= 200 and row["风险等级"] != "高":
            row["建议操作"] = "✅ 可以测试"
        elif p.historical_sold >= 50:
            row["建议操作"] = "👀 观察"
        elif p.historical_sold >= 10:
            row["建议操作"] = "📋 缺销量数据"
        else:
            row["建议操作"] = "⬇️ 数据不足"
        
        results.append(row)
    
    # 按毛利率排序
    results.sort(key=lambda r: r["毛利率(%)"], reverse=True)
    return results


def export_excel(rows: list[dict], filename: str = None):
    """导出 Excel"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"选品分析_{ts}.xlsx"
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # 回退到 CSV
        csv_path = os.path.join(REPORTS_DIR, filename.replace(".xlsx", ".csv"))
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=rows[0].keys())
            w.writeheader()
            w.writerows(rows)
        return csv_path
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "选品分析"
    
    headers = list(rows[0].keys())
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin = Border(*(Side(style="thin"),)*4)
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = Alignment(horizontal="center"); c.border = thin
    
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin
            # 条件颜色
            if h == "建议操作":
                if "✅" in str(val):
                    cell.fill = green_fill
                elif "❌" in str(val):
                    cell.fill = red_fill
                elif "👀" in str(val):
                    cell.fill = yellow_fill
    
    widths = [45, 10, 10, 8, 6, 8, 10, 10, 8, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    
    path = os.path.join(REPORTS_DIR, filename)
    wb.save(path)
    return path


# ============================================================
# 示例数据 & 手动输入
# ============================================================

def load_sample_data() -> list[Product]:
    """加载内置示例数据（桌面收纳类目）"""
    return [
        Product("北欧风桌面收纳盒 创意整理盒 办公桌置物架", 199, 12, 850, 4.8, 320),
        Product("透明亚克力化妆品收纳盒 防尘护肤收纳", 299, 18, 1200, 4.7, 560),
        Product("抽屉式桌面收纳柜 A4文件杂物整理箱", 399, 25, 650, 4.5, 280),
        Product("多功能笔筒收纳 旋转桌面文具架", 159, 8, 420, 4.6, 180),
        Product("桌面手机支架 可调节追剧神器", 89, 5, 3200, 4.3, 1500, shipping_twd=40),
        Product("分层桌面书架 学生宿舍桌上置物架", 259, 16, 380, 4.4, 165),
        Product("数据线收纳盒 绕线器 桌面理线神器", 69, 4, 5600, 4.5, 2400, shipping_twd=30),
        Product("桌面垃圾桶 迷你按压式 办公桌垃圾盒", 79, 5, 2100, 4.6, 890, shipping_twd=35),
        Product("电脑增高架 显示器支架 桌面收纳架", 349, 22, 980, 4.7, 430),
        Product("毛毡桌面收纳 创意笔筒平板支架", 129, 8, 340, 4.3, 120),
        Product("可伸缩桌面书架 学生宿舍整理架", 189, 11, 520, 4.5, 230),
        Product("魔方插座 桌面USB快充排插", 249, 15, 1800, 4.4, 890),
        Product("桌面小抽屉储物盒 内衣收纳 迷你柜", 139, 8, 670, 4.6, 290),
        Product("亚克力口红收纳架 梳妆台化妆刷桶", 179, 10, 890, 4.7, 410),
        Product("桌面理线器 数据线固定夹 3个装", 49, 3, 7800, 4.2, 3400, shipping_twd=25),
    ]


def load_from_csv(path: str) -> list[Product]:
    """从 CSV 文件加载数据"""
    products = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            p = Product(
                title=row.get("标题", row.get("title", "")),
                price_twd=float(row.get("售价", row.get("price_twd", 0))),
                price_cny=float(row.get("批发价", row.get("price_cny", 0))),
                historical_sold=int(row.get("月销", row.get("historical_sold", 0))),
                rating=float(row.get("评分", row.get("rating", 0))),
                review_count=int(row.get("评价数", row.get("review_count", 0))),
            )
            products.append(p)
    return products


# ============================================================
# 1688 爬虫（轻量版）
# ============================================================

def crawl_1688_price(keyword: str = "桌面收纳"):
    """快速获取 1688 价格参考（用于比价）"""
    import requests
    from bs4 import BeautifulSoup
    
    url = f"https://s.1688.com/selloffer/offer_search.htm?keywords={keyword}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html",
    }
    try:
        r = requests.get(url, headers=headers, proxies={"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"}, timeout=15)
        if r.status_code != 200:
            return None
        
        soup = BeautifulSoup(r.text, "html.parser")
        items = soup.select(".offer-list-row .offer-item") or soup.select("[data-offer-id]")
        
        prices = []
        for item in items[:20]:
            price_el = item.select_one(".price") or item.select_one("[data-price]")
            title_el = item.select_one(".title") or item.select_one("a[title]")
            if price_el:
                price_text = price_el.get_text(strip=True)
                import re
                match = re.search(r"[\d.]+", price_text)
                if match:
                    prices.append(float(match.group()))
        
        if prices:
            return {
                "avg_price": round(sum(prices) / len(prices), 2),
                "min_price": min(prices),
                "max_price": max(prices),
                "samples": len(prices),
                "keyword": keyword,
            }
    except Exception as e:
        return {"error": str(e)}
    return None


# ============================================================
# 入口
# ============================================================

def main():
    print(f"\n📦 Shopee 选品分析工具 v1.0")
    print(f"{'='*50}")
    
    # 加载数据
    if len(sys.argv) > 1:
        path = sys.argv[1]
        if os.path.exists(path):
            products = load_from_csv(path)
            print(f"📂 从文件加载: {path}")
        else:
            print(f"❌ 文件不存在: {path}")
            sys.exit(1)
    else:
        products = load_sample_data()
        print(f"📋 使用示例数据（桌面收纳）")
        print(f"   💡 运行: python3 main.py <your-data.csv>")
        print(f"   CSV格式: 标题,售价,批发价,月销,评分,评价数")
    
    print(f"   📊 共 {len(products)} 个商品\n")
    
    # 尝试获取 1688 价格参考（可跳过）
    # 1688 有反爬，暂时用估算
    print("   💡 批发价(CNY) 填你从 1688 找到的进货价")
    print()
    
    # 分析
    results = analyze_products(products)
    
    # 输出摘要
    print(f"{'='*50}")
    print(f"📋 分析结果（按毛利率排序）")
    print(f"{'='*50}")
    # 中英文混排对齐：用 pad_disp 确保中文双宽字符对齐
    print(f"  {pad_disp('商品', 30)} {'售价':>8} {'毛利':>7} {'月销':>6} {'评分':>4}  建议")
    print(f"  {'─'*66}")
    
    good = bad = 0
    for r in results:
        title = r["标题"][:16] + ".." if disp_width(r["标题"]) > 32 else r["标题"]
        print(f"  {pad_disp(title, 30)} {r['售价(TWD)']:>8.0f} {r['毛利率(%)']:>7}% {r['月销']:>6} {r['评分']:>4.1f} {r['建议操作']}")
        if "✅" in r["建议操作"]:
            good += 1
        elif "❌" in r["建议操作"]:
            bad += 1
    
    print(f"\n📊 统计: {good} 个推荐 ✅ | {bad} 个不推荐 ❌ | {len(results)-good-bad} 个待观察 👀")
    
    # 导出
    xlsx_path = export_excel(results)
    print(f"\n📊 已导出: {xlsx_path}")
    print(f"✅ 完成！")


if __name__ == "__main__":
    main()
